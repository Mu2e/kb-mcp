"""Excel parser for extracting text from XLSX files."""

import logging
from pathlib import Path

from .parser_base import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel spreadsheets (XLSX)."""

    def extract_text(self) -> str:
        """Extract text from Excel spreadsheet."""
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "openpyxl not installed. Install with: pip install openpyxl"
            )
            return ""

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error extracting text from Excel {self.file_path}: {e}")
            return ""

